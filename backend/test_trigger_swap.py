import os
import sys
import datetime
import openpyxl
import pandas as pd

# Add backend to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from backend.models import DailyMenu, MenuDish, MenuTable
from backend.menu_generator import generate_kondate_excel, save_menu_master

def test_trigger_swap():
    print("Testing Trigger-Based Menu Swapping...")
    
    # 1. Mock Menu Master
    today = datetime.date.today()
    
    # Base Menu (Usually for this date)
    base_dishes = [MenuDish(dish_name=f"Base Dish {i}") for i in range(6)]
    base_menu = DailyMenu(date=today, dishes=base_dishes)
    
    # Special Block (identified by trigger "カレー" in Master)
    special_dishes = [MenuDish(dish_name=f"Curry Part {i}") for i in range(6)]
    special_menu = DailyMenu(meal_type="カレー", dishes=special_dishes)
    
    master = MenuTable(
        year=today.year,
        month=today.month,
        base_menus={today: base_menu},
        special_menus={"カレー": special_menu},
        allergy_menus={}
    )
    
    # Save Master
    save_menu_master(master)
    
    # 2. Case A: Order is "通常"
    options_normal = {
        "kindergarten_name": "Test Kindy",
        "orders": [{"date": today.strftime("%Y-%m-%d"), "meal_type": "通常"}]
    }
    path_normal = generate_kondate_excel("TEST_A", today.year, today.month, options_normal)
    
    wb_n = openpyxl.load_workbook(path_normal)
    ws_n = wb_n.active
    # Row 8 Col 2 should be "Base Dish 0"
    val_n = ws_n.cell(row=8, column=2).value
    if val_n == "Base Dish 0":
        print("✅ Case A (Normal): Correctly kept base menu.")
    else:
        print(f"❌ Case A (Normal) Failed: {val_n}")

    # 3. Case B: Order is "カレー" (Exact Match)
    options_curry = {
        "kindergarten_name": "Test Kindy",
        "orders": [{"date": today.strftime("%Y-%m-%d"), "meal_type": "カレー"}]
    }
    path_curry = generate_kondate_excel("TEST_B", today.year, today.month, options_curry)
    wb_c = openpyxl.load_workbook(path_curry)
    ws_c = wb_c.active
    # Row 8 Col 2 should be "Curry Part 0"
    val_c = ws_c.cell(row=8, column=2).value
    if val_c == "Curry Part 0":
        print("✅ Case B (Curry Exact): Correctly swapped to Curry menu.")
    else:
        print(f"❌ Case B (Curry Exact) Failed: {val_c}")

    # 4. Case C: Order is "カレ" (Fuzzy Match)
    options_fuzzy = {
        "kindergarten_name": "Test Kindy",
        "orders": [{"date": today.strftime("%Y-%m-%d"), "meal_type": "カレ"}]
    }
    path_fuzzy = generate_kondate_excel("TEST_C", today.year, today.month, options_fuzzy)
    wb_f = openpyxl.load_workbook(path_fuzzy)
    ws_f = wb_f.active
    # Row 8 Col 2 should be "Curry Part 0"
    val_f = ws_f.cell(row=8, column=2).value
    if val_f == "Curry Part 0":
        print("✅ Case C (Curry Fuzzy): Correctly swapped to Curry menu via fuzzy match.")
    else:
        print(f"❌ Case C (Curry Fuzzy) Failed: {val_f}")

if __name__ == "__main__":
    test_trigger_swap()
