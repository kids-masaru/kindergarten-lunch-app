import os
import sys
import datetime
import openpyxl

# Add backend to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from backend.models import DailyMenu, MenuDish, MenuTable
from backend.menu_generator import generate_kondate_excel, save_menu_master

def test_structure():
    print("Testing 6-row block structure and dynamic resolution...")
    
    # 1. Create a mock DailyMenu with exactly 6 dishes
    dishes = [
        MenuDish(dish_name="ごはん", ingredients_red="米", ingredients_yellow="なし", ingredients_green="なし", seasoning="なし", remarks=""),
        MenuDish(dish_name="味噌汁", ingredients_red="味噌", ingredients_yellow="わかめ", ingredients_green="なし", seasoning="だし", remarks=""),
        MenuDish(dish_name="焼肉", ingredients_red="豚肉", ingredients_yellow="油", ingredients_green="キャベツ", seasoning="醤油", remarks=""),
        MenuDish(dish_name="", ingredients_red=None, ingredients_yellow=None, ingredients_green=None, seasoning=None, remarks=None), # Empty row
        MenuDish(dish_name="", ingredients_red=None, ingredients_yellow=None, ingredients_green=None, seasoning=None, remarks=None), # Empty row
        MenuDish(dish_name="デザート", ingredients_red="いちご", ingredients_yellow=None, ingredients_green=None, seasoning=None, remarks="甘い")
    ]
    
    today = datetime.date.today()
    base_menu = DailyMenu(
        date=today,
        meal_type="通常",
        dishes=dishes,
        total_energy=500.0,
        total_protein=20.0,
        total_lipid=15.0
    )
    
    # 2. Create Special Menu
    special_dishes = [MenuDish(dish_name="カレーライス") for _ in range(6)]
    special_menu = DailyMenu(date=today, meal_type="カレー", dishes=special_dishes)
    
    master = MenuTable(
        year=today.year,
        month=today.month,
        base_menus={today: base_menu},
        allergy_menus={},
        special_menus={"カレー": special_menu},
        kindergarten_sheets={}
    )
    
    # Save Master
    save_menu_master(master)
    
    # 3. Test Generation for "usually"
    options_normal = {
        "kindergarten_name": "テスト幼稚園",
        "orders": [{"date": today.strftime("%Y-%m-%d"), "meal_type": "通常"}]
    }
    
    print("Generating Normal Menu...")
    path_normal = generate_kondate_excel("test_kind", today.year, today.month, options_normal)
    print(f"Generated: {path_normal}")
    
    # Verify content
    wb = openpyxl.load_workbook(path_normal)
    ws = wb.active
    
    # Check if Row 8 Col B is "ごはん"
    # Row 8 is current_row. Col 2 is B.
    if ws.cell(row=8, column=2).value == "ごはん":
        print("✅ Row 1 dish name correct")
    else:
        print(f"❌ Row 1 dish name mismatch: {ws.cell(row=8, column=2).value}")

    # Check color-coded ingredients (C=3, D=4, E=5)
    if ws.cell(row=8, column=3).value == "米":
        print("✅ Red ingredient correct")
    
    # Check Row 6 of block (Row 13)
    if ws.cell(row=13, column=2).value == "デザート":
        print("✅ Row 6 dish name correct")
    else:
        print(f"❌ Row 6 dish name mismatch: {ws.cell(row=13, column=2).value}")

    # 4. Test Generation for "Curry"
    options_curry = {
        "kindergarten_name": "テスト幼稚園",
        "orders": [{"date": today.strftime("%Y-%m-%d"), "meal_type": "カレー"}]
    }
    print("Generating Curry Menu...")
    path_curry = generate_kondate_excel("test_kind", today.year, today.month, options_curry)
    wb_curry = openpyxl.load_workbook(path_curry)
    ws_curry = wb_curry.active
    if ws_curry.cell(row=8, column=2).value == "カレーライス":
        print("✅ Dynamic selection (Curry) correct")
    else:
        print(f"❌ Dynamic selection (Curry) failed: {ws_curry.cell(row=8, column=2).value}")

if __name__ == "__main__":
    test_structure()
