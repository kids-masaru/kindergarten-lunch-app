import pandas as pd
import datetime
import os
import json
from typing import Dict, List, Optional
from backend.models import MenuTable, DailyMenu, MenuDish, normalize_key
from backend.drive import upload_file_to_drive, download_file_from_drive

# Directory to store Menu Masters (JSON) and Generated Files
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
os.makedirs(DATA_DIR, exist_ok=True)

def save_menu_master(table: MenuTable):
    """Saves the parsed MenuTable to a JSON file and uploads to Drive."""
    filename = f"menu_master_{table.year}_{table.month}.json"
    filepath = os.path.join(DATA_DIR, filename)
    
    # Serialization logic handling Pydantic v1/v2 compatibility
    with open(filepath, 'w', encoding='utf-8') as f:
        # Try Pydantic v2
        if hasattr(table, 'model_dump_json'):
             # v2
             f.write(table.model_dump_json())
        else:
             # v1 - .json() should work, but ensure_ascii might be an issue in some versions?
             # Let's try standard json dumps on .dict() with date handler
             def json_serial(obj):
                if isinstance(obj, (datetime.date, datetime.datetime)):
                    return obj.isoformat()
                raise TypeError (f"Type {type(obj)} not serializable")
                
             json.dump(table.dict(), f, default=json_serial, ensure_ascii=False, indent=2)
             
    # Upload to Drive
    # Upload to Drive
    print(f"Uploading Master to Drive: {filename}")
    try:
        upload_file_to_drive(filepath, filename, mime_type='application/json')
    except Exception as e:
        print(f"[WARNING] Failed to upload Master to Drive (likely quota issue): {e}")
        # Continue execution even if backup fails
             
    return filepath

def load_menu_master(year: int, month: int) -> Optional[MenuTable]:
    """Loads a MenuTable from JSON (Downloads from Drive if needed)."""
    filename = f"menu_master_{year}_{month}.json"
    filepath = os.path.join(DATA_DIR, filename)
    
    if not os.path.exists(filepath):
        # Try to download from Drive
        print(f"Local master not found. Checking Drive for: {filename}")
        success = download_file_from_drive(filename, filepath)
        if not success:
            return None
        
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
        return MenuTable(**data)

import openpyxl
from openpyxl.utils import get_column_letter

# Paths
BACKEND_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BACKEND_DIR, '..', 'data')
os.makedirs(DATA_DIR, exist_ok=True)
TEMPLATE_FILE = os.path.join(BACKEND_DIR, 'data', 'template.xlsx')

def generate_kondate_excel(kindergarten_id: str, year: int, month: int, options: Dict) -> str:
    """
    Generates the Kondate Excel for a specific kindergarten using template.xlsx.
    """
    # 1. Load Master
    master = load_menu_master(year, month)
    if not master:
        raise ValueError(f"Menu Master not found for {year}-{month}")

    # 2. Resolve Daily Menus
    daily_schedule = resolve_daily_menus(master, kindergarten_id, options)
    
    # 3. Load Template
    if not os.path.exists(TEMPLATE_FILE):
        # Fallback to simple excel if template missing
        print(f"[WARNING] Template not found at {TEMPLATE_FILE}. Generating simple file.")
        return generate_simple_excel(daily_schedule, kindergarten_id, year, month)

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb.active # Use the first sheet
    
    # 4. Fill Logic
    # Start Left: Row 8. 
    # Wrap after Row 73 -> Move to Col J (index 10), Row 5.
    
    current_row = 8
    current_col_offset = 0 # 0 for A-H, 9 for J-Q (J is index 10, but offset from A is 9)
    
    sorted_dates = sorted(daily_schedule.keys())
    
    for d in sorted_dates:
        menu = daily_schedule[d]
        
        # Check for wrap-around
        if current_row + 5 > 73 and current_col_offset == 0:
            current_row = 5
            current_col_offset = 9 # Moves window from A-H to J-Q (J is 10th col)
        
        # Fill Date (Col A + offset)
        # ws.cell(row=current_row, column=1 + current_col_offset, value=d.day) # Date
        # ws.cell(row=current_row + 1, column=1 + current_col_offset, value=d.strftime("%a")) # Day
        # Actually user said A7 is date, A8 is day in the SOURCE. 
        # For TEMPLATE, let's assume Row 1 of block is date, Row 2 is day.
        ws.cell(row=current_row, column=1 + current_col_offset, value=d.day)
        
        # Fill Dishes (Strictly 6)
        for i in range(6):
            r = current_row + i
            dish = menu.dishes[i] if i < len(menu.dishes) else MenuDish(dish_name="")
            
            # B: Menu (offset 1)
            ws.cell(row=r, column=2 + current_col_offset, value=dish.dish_name)
            # C, D, E: Ingredients (Offset 2, 3, 4)
            ws.cell(row=r, column=3 + current_col_offset, value=dish.ingredients_red)
            ws.cell(row=r, column=4 + current_col_offset, value=dish.ingredients_yellow)
            ws.cell(row=r, column=5 + current_col_offset, value=dish.ingredients_green)
            # F: Seasoning (offset 5)
            ws.cell(row=r, column=6 + current_col_offset, value=dish.seasoning)
            # H: Remarks (offset 7)
            ws.cell(row=r, column=8 + current_col_offset, value=dish.remarks)
            
        # Fill Nutrition totals in Col G (offset 6)
        # G2: Energy (row+1), G4: Protein (row+3), G6: Lipid (row+5)
        if menu.total_energy:
            ws.cell(row=current_row + 1, column=7 + current_col_offset, value=menu.total_energy)
        if menu.total_protein:
            ws.cell(row=current_row + 3, column=7 + current_col_offset, value=menu.total_protein)
        if menu.total_lipid:
            ws.cell(row=current_row + 5, column=7 + current_col_offset, value=menu.total_lipid)

        current_row += 6 # Next block

    # Update Title (Usually somewhere in the top)
    ws['A1'] = f"{year}年 {month}月 献立表"
    ws['E1'] = options.get('kindergarten_name', '幼稚園')

    # 5. Save
    output_filename = f"kondate_{kindergarten_id}_{year}_{month}.xlsx"
    output_path = os.path.join(DATA_DIR, output_filename)
    wb.save(output_path)
    
    return output_path

def generate_simple_excel(schedule: Dict, kid: str, y: int, m: int) -> str:
    """Fallback generator that preserves the 6-row block structure."""
    all_rows = []
    # Header
    all_rows.append(["Date", "Day", "Dish Name", "Red", "Yellow", "Green", "Seasoning", "Nutrition", "Remarks"])
    
    for d in sorted(schedule.keys()):
        menu = schedule[d]
        for i in range(6):
            dish = menu.dishes[i] if i < len(menu.dishes) else MenuDish(dish_name="")
            row = [
                d.strftime("%Y-%m-%d") if i == 0 else "", 
                d.strftime("%a") if i == 0 else "",
                dish.dish_name,
                dish.ingredients_red,
                dish.ingredients_yellow,
                dish.ingredients_green,
                dish.seasoning,
                # Simple nutrition list
                f"E:{menu.total_energy}" if i == 1 else (f"P:{menu.total_protein}" if i == 3 else (f"L:{menu.total_lipid}" if i == 5 else "")),
                dish.remarks
            ]
            all_rows.append(row)
            
    df = pd.DataFrame(all_rows[1:], columns=all_rows[0])
    path = os.path.join(DATA_DIR, f"kondate_{kid}_{y}_{m}_fallback.xlsx")
    df.to_excel(path, index=False)
    return path

import difflib

def resolve_daily_menus(master: MenuTable, kindergarten_id: str, options: Dict) -> Dict[datetime.date, DailyMenu]:
    """
    Core Logic: Merges Base + Special + Allergy + Options to produce final daily menus.
    Uses trigger-based swapping for special menus.
    """
    schedule = {}
    
    # Orders for this month might contain the menu types selected by the user
    orders = options.get('orders', [])
    
    # Map Orders by Date (YYYY-MM-DD -> meal_type)
    date_to_meal_type = {}
    for o in orders:
        if isinstance(o, dict):
            date_to_meal_type[o.get('date')] = o.get('meal_type')
        else:
            date_to_meal_type[o.date] = o.meal_type
    
    # Prepared list of special menu keys for fuzzy matching
    special_keys = list(master.special_menus.keys())

    # 2. Iterate days
    for date, base_menu in master.base_menus.items():
        final_menu = base_menu
        date_str = date.strftime("%Y-%m-%d")
        
        # Determine Meal Type from Order (Default to "通常")
        meal_type = date_to_meal_type.get(date_str, "通常")
        
        # --- Trigger Swapping Logic ---
        # 1. Exact Match
        if meal_type in master.special_menus:
            final_menu = master.special_menus[meal_type]
            final_menu.date = date
        # 2. Allergy/Catering specific (Legacy "配膳" fallback)
        elif meal_type == "配膳" and date in master.allergy_menus:
            final_menu = master.allergy_menus[date]
        # 3. Fuzzy Match (if not "通常" or "飯なし")
        elif meal_type not in ["通常", "飯なし"]:
            matches = difflib.get_close_matches(meal_type, special_keys, n=1, cutoff=0.5)
            if matches:
                match_key = matches[0]
                print(f"Fuzzy matched '{meal_type}' to '{match_key}'")
                final_menu = master.special_menus[match_key]
                final_menu.date = date

        # Check kindergarten-specific sheet overrides (Sheet matching)
        kind_name = options.get('kindergarten_name')
        if kind_name and kind_name in master.kindergarten_sheets:
            kind_overrides = master.kindergarten_sheets[kind_name]
            if date in kind_overrides:
                final_menu = kind_overrides[date]

        schedule[date] = final_menu
        
    return schedule
